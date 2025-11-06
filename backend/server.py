from fastapi import FastAPI, APIRouter, HTTPException, status, UploadFile, File
from fastapi import Request
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
from datetime import datetime, timezone, timedelta
import bcrypt
from contextlib import asynccontextmanager
import io
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import asyncio
import openpyxl
import json
import stripe

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from .database import db, client

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Sentry (if configured) as early as possible to capture import-time errors.
try:
    from .monitoring import init_sentry as _init_sentry
except Exception:
    # fallback to top-level import if package context differs
    try:
        from monitoring import init_sentry as _init_sentry
    except Exception:
        _init_sentry = None

# Minimal FastAPI app object to ensure decorators defined later don't fail during import.
# A proper lifespan/context is configured further down; this early creation prevents
# NameError during module import when decorators like @app.get are evaluated.
try:
    from fastapi import FastAPI as _FastAPI
    app = _FastAPI()
    if _init_sentry:
        try:
            _init_sentry(app)
        except Exception:
            logger.exception("Sentry init failed during early app creation")
except Exception:
    app = None  # fallback; import errors will surface later

# Initialize structured logging and request-id middleware
try:
    from .logging_config import init_logging, RequestIDMiddleware
except Exception:
    try:
        from logging_config import init_logging, RequestIDMiddleware
    except Exception:
        init_logging = None
        RequestIDMiddleware = None

if init_logging:
    try:
        init_logging()
    except Exception:
        logger.exception("Failed to initialize structured logging")

if app and RequestIDMiddleware:
    try:
        app.add_middleware(RequestIDMiddleware)
    except Exception:
        logger.exception("Failed to add RequestIDMiddleware")

# Lifespan context manager for proper shutdown
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Lifespan startup/shutdown hook — keep lightweight and avoid expensive work here.
    # Any heavy precomputation (e.g. salary reports) should be queued via background jobs
    # or triggered by explicit endpoints. Do not reference request-specific variables
    # such as `month` in this module-level lifespan.
    try:
        # Ensure POS collections and recommended indexes exist. Non-destructive.
        try:
            from .pos_collections import ensure_pos_collections
        except Exception:
            try:
                from pos_collections import ensure_pos_collections
            except Exception:
                ensure_pos_collections = None

        if ensure_pos_collections:
            try:
                # run but don't block startup if it fails
                await ensure_pos_collections(db)
            except Exception:
                logger.exception('pos_collections.ensure_pos_collections failed')
        yield
    finally:
        # perform any graceful shutdown tasks here if needed
        try:
            # Motor's AsyncIOMotorClient.close is synchronous; call without await
            client.close()
        except Exception:
            pass

# Add CORS middleware BEFORE routers so browser requests from the frontend are allowed.
if app:
    try:
        app.add_middleware(
            CORSMiddleware,
            allow_origins=os.environ.get('CORS_ORIGINS', 'https://mevcut-appv1.vercel.app,*').split(','),
            allow_methods=["*"],
            allow_headers=["*"],
            # Do not allow credentials with a wildcard origin — keep it False for public deploys
            allow_credentials=False,
        )
    except Exception:
        logger.exception("Failed to add CORS middleware")

# Create router after middleware
api_router = APIRouter(prefix="/api")

# If a frontend build exists next to this backend, mount it so the backend
# can serve the static assets (temporary fallback while hosting is fixed).
FRONTEND_BUILD_DIR = ROOT_DIR.parent / "frontend" / "build"
if FRONTEND_BUILD_DIR.exists():
    try:
        static_dir = FRONTEND_BUILD_DIR / "static"
        if static_dir.exists():
            app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")
            logger.info("Mounted frontend static directory at /static from %s", static_dir)
        else:
            logger.info("Frontend build exists but no static/ subfolder found at %s", FRONTEND_BUILD_DIR)
    except Exception:
        logger.exception("Failed to mount frontend static files (temporary SPA fallback)")

from .models import *
from .routers import employees, tasks, shifts, stock

# ==================== HELPER FUNCTIONS ====================


# ==================== ROUTES ====================

# Health Check
@app.get("/")
async def root():
    return {
        "status": "ok",
        "message": "API is running",
        "login_methods": {
            "method_1_email": {
                "endpoint": "/api/login",
                "email": "admin@example.com",
                "password": "admin123"
            },
            "method_2_employee_id": {
                "endpoint": "/api/login",
                "employee_id": "1000",
                "password": "admin123",
                "company_id": 1
            },
            "method_3_pin": {
                "endpoint": "/api/pin-login",
                "employee_id": "1000",
                "pin": "1234",
                "note": "Admin PIN: 1234, Arda PIN: 2024, Others: use employee_id as PIN"
            }
        },
        "test_users": {
            "admin": {"employee_id": "1000", "pin": "1234", "password": "admin123"},
            "arda": {"employee_id": "2001", "pin": "2024", "password": "arda2024"}
        },
        "test_endpoints": [
            "/api/test-login",
            "/api/login-test?employee_id=1000&password=admin123",
            "/api/check-user/1000",
            "/api/ensure-admin"
        ]
    }

# Test login endpoint
@api_router.get("/test-login")
async def test_login():
    """Test endpoint to check available users"""
    users = await db.employees.find({}, {"password": 0}).to_list(None)
    return {
        "total_users": len(users),
        "users": [
            {
                "id": u.get("id"),
                "employee_id": u.get("employee_id"),
                "email": u.get("email"),
                "name": f"{u.get('ad')} {u.get('soyad')}",
                "rol": u.get("rol"),
                "has_password": bool(u.get("password"))
            }
            for u in users
        ],
        "login_info": "Use /api/login with employee_id or email + password"
    }

app.include_router(employees.router, prefix="/api", tags=["employees"])
app.include_router(tasks.router, prefix="/api", tags=["tasks"])

# Role Routes
@api_router.get("/roles", response_model=List[Role])
async def get_roles():
    roles = await db.roles.find().to_list(None)
    return roles

@api_router.put("/roles/{role_id}", response_model=Role)
async def update_role(role_id: str, role_update: RoleUpdate):
    result = await db.roles.update_one(
        {"id": role_id},
        {"$set": {"permissions": role_update.permissions.dict()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Role not found")
    
    updated_role = await db.roles.find_one({"id": role_id})
    return updated_role

# Shift Type Routes
@api_router.get("/shift-types", response_model=List[ShiftType])
async def get_shift_types():
    shift_types = await db.shift_types.find().to_list(None)
    return shift_types

@api_router.post("/shift-types", response_model=ShiftType)
async def create_shift_type(shift_type: ShiftTypeCreate):
    import uuid
    new_shift_type = {
        "id": str(uuid.uuid4()),
        **shift_type.dict()
    }
    await db.shift_types.insert_one(new_shift_type)
    return new_shift_type

@api_router.delete("/shift-types/{shift_type_id}")
async def delete_shift_type(shift_type_id: str):
    result = await db.shift_types.delete_one({"id": shift_type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Shift type not found")
    return {"message": "Shift type deleted successfully"}

# Attendance Routes
@api_router.get("/attendance", response_model=List[Attendance])
async def get_attendance(company_id: int = 1, date: Optional[str] = None):
    query = {"company_id": company_id}
    if date:
        query["tarih"] = date
    attendance = await db.attendance.find(query).to_list(None)
    return attendance

@api_router.post("/attendance/check-in")
async def check_in(check_in_data: AttendanceCheckIn):
    # Find employee
    employee = await db.employees.find_one({
        "company_id": check_in_data.company_id,
        "employee_id": check_in_data.employee_id
    })
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Check if already checked in today
    existing = await db.attendance.find_one({
        "company_id": check_in_data.company_id,
        "employee_id": check_in_data.employee_id,
        "tarih": today
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Already checked in today")
    
    # Create attendance record
    next_id = await get_next_id("attendance")
    attendance_record = {
        "id": next_id,
        "company_id": check_in_data.company_id,
        "employee_id": check_in_data.employee_id,
        "ad": employee["ad"],
        "soyad": employee["soyad"],
        "tarih": today,
        "giris_saati": datetime.now(timezone.utc).isoformat(),
        "cikis_saati": None,
        "calisilan_saat": 0,
        "status": "giris"
    }
    
    await db.attendance.insert_one(attendance_record)
    return {"message": "Check-in successful", "time": attendance_record["giris_saati"]}

@api_router.post("/attendance/check-out")
async def check_out(check_out_data: AttendanceCheckOut):
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Find today's attendance record
    attendance = await db.attendance.find_one({
        "company_id": check_out_data.company_id,
        "employee_id": check_out_data.employee_id,
        "tarih": today,
        "status": "giris"
    })
    
    if not attendance:
        raise HTTPException(status_code=404, detail="No check-in found for today")
    
    # Calculate worked hours
    check_out_time = datetime.now(timezone.utc)
    check_in_time = datetime.fromisoformat(attendance["giris_saati"])
    worked_hours = (check_out_time - check_in_time).total_seconds() / 3600
    
    # Update attendance record
    await db.attendance.update_one(
        {"id": attendance["id"]},
        {
            "$set": {
                "cikis_saati": check_out_time.isoformat(),
                "calisilan_saat": round(worked_hours, 2),
                "status": "cikis"
            }
        }
    )
    
    return {
        "message": "Check-out successful",
        "time": check_out_time.isoformat(),
        "worked_hours": round(worked_hours, 2)
    }

app.include_router(shifts.router, prefix="/api", tags=["shifts"])

# Admin check/create endpoint
@api_router.post("/ensure-admin")
async def ensure_admin():
    """Ensure admin user exists with admin@example.com"""
    # Check if admin exists
    admin_user = await db.employees.find_one({"email": "admin@example.com"})
    
    if admin_user:
        return {"message": "Admin user already exists", "email": "admin@example.com"}
    
    # Check if admin role exists
    admin_role = await db.roles.find_one({"id": "admin"})
    if not admin_role:
        # Create admin role
        admin_role = {
            "id": "admin",
            "name": "Yönetici",
            "permissions": {
                "view_dashboard": True,
                "view_tasks": True,
                "assign_tasks": True,
                "rate_tasks": True,
                "manage_shifts": True,
                "manage_leave": True,
                "view_salary": True,
                "manage_roles": True,
                "manage_shifts_types": True,
                "edit_employees": True,
                "can_view_stock": True,
                "can_add_stock_unit": True,
                "can_delete_stock_unit": True,
                "can_add_stock_product": True,
                "can_edit_stock_product": True,
                "can_delete_stock_product": True,
                "can_perform_stock_count": True,
                "can_manage_categories": True
            }
        }
        await db.roles.insert_one(admin_role)
    
    # Check if company exists
    company = await db.companies.find_one({"id": 1})
    if not company:
        company = {
            "id": 1,
            "name": "Demo Şirket",
            "domain": "example.com",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.companies.insert_one(company)
    
    # Create admin user
    next_id = await get_next_id("employees")
    admin_user = {
        "id": next_id,
        "company_id": 1,
        "ad": "Admin",
        "soyad": "User",
        "pozisyon": "Sistem Yöneticisi",
        "maas_tabani": 50000,
        "rol": "admin",
        "email": "admin@example.com",
        "employee_id": "1000",
        "password": bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    }
    await db.employees.insert_one(admin_user)
    
    return {
        "message": "Admin user created successfully",
        "email": "admin@example.com",
        "password": "admin123",
        "note": "Please change the password after first login"
    }

# Migration endpoint - Add passwords to existing users
@api_router.post("/migrate-passwords")
async def migrate_passwords():
    """Add default passwords to existing users without passwords"""
    updated_count = 0
    
    # Default passwords for known users
    default_passwords = {
        "1000": "admin123",
        "1001": "mehmet123",
        "1002": "zeynep123",
        "1003": "ayse123", 
        "1004": "ali123",
        "2001": "arda2024"
    }
    
    # Get all employees
    employees = await db.employees.find({}).to_list(None)
    
    for employee in employees:
        if not employee.get("password"):
            employee_id = employee.get("employee_id")
            # Use default password based on employee_id or generate one
            if employee_id in default_passwords:
                password = default_passwords[employee_id]
            else:
                password = f"user{employee_id}"
            
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            await db.employees.update_one(
                {"id": employee["id"]},
                {"$set": {"password": hashed_password}}
            )
            updated_count += 1
            logger.info(f"Password set for {employee['ad']} {employee['soyad']} (ID: {employee_id})")
    
    return {
        "message": f"Passwords updated for {updated_count} users",
        "default_passwords": {
            "admin (1000)": "admin123",
            "arda (2001)": "arda2024",
            "others": "Check logs or use default pattern"
        }
    }


# Admin password reset endpoint (gated by env var)
@api_router.post("/reset-admin-password")
async def reset_admin_password(payload: Dict[str, str]):
    """Reset admin@example.com's password to a provided value (or admin123 by default).

    This endpoint is gated by the ALLOW_ADMIN_RESET environment variable to avoid accidental use.
    Request JSON: { "password": "newpass" }
    """
    if os.environ.get("ALLOW_ADMIN_RESET", "false").lower() != "true":
        raise HTTPException(status_code=403, detail="Admin password reset not allowed")

    new_password = payload.get("password") or "admin123"
    hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    result = await db.employees.update_one({"email": "admin@example.com"}, {"$set": {"password": hashed}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Admin user not found")

    return {"message": "Admin password updated", "email": "admin@example.com"}


# Compatibility aliases for frontend endpoints that use dashed paths
# These simply proxy to the canonical /stok/... endpoints above
@api_router.get("/stok-birim")
async def get_stok_birim_alias(company_id: int = 1):
    return await get_stok_birimleri(company_id)

@api_router.post("/stok-birim")
async def post_stok_birim_alias(birim: StokBirimCreate):
    return await create_stok_birim(birim)

@api_router.delete("/stok-birim/{birim_id}")
async def delete_stok_birim_alias(birim_id: int):
    return await delete_stok_birim(birim_id)

@api_router.get("/stok-kategori")
async def get_stok_kategori_alias(company_id: int = 1):
    return await get_stok_kategorileri(company_id)

@api_router.post("/stok-kategori")
async def post_stok_kategori_alias(kategori: StokKategoriCreate):
    return await create_stok_kategori(kategori)

@api_router.put("/stok-kategori/{kategori_id}")
async def put_stok_kategori_alias(kategori_id: int, kategori: StokKategoriCreate):
    return await update_stok_kategori(kategori_id, kategori)

@api_router.get("/stok-urun")
async def get_stok_urun_alias(company_id: int = 1):
    return await get_stok_urunleri(company_id)

@api_router.post("/stok-urun")
async def post_stok_urun_alias(urun: StokUrunCreate):
    return await create_stok_urun(urun)

@api_router.put("/stok-urun/{urun_id}")
async def put_stok_urun_alias(urun_id: int, urun_update: StokUrunUpdate):
    return await update_stok_urun(urun_id, urun_update)

@api_router.delete("/stok-urun/{urun_id}")
async def delete_stok_urun_alias(urun_id: int):
    return await delete_stok_urun(urun_id)

@api_router.get("/stok-sayim/son-durum")
async def stok_son_durum_alias(company_id: int = 1):
    # Return last counts per product - reuse get_stok_sayimlari
    sayimlar = await get_stok_sayimlari(company_id)
    return sayimlar

@api_router.post("/stok-sayim")
async def post_stok_sayim_alias(sayim: StokSayimCreate):
    return await create_stok_sayim(sayim)

app.include_router(stock.router, prefix="/api", tags=["stock"])

# Seed data endpoint
@api_router.post("/seed-data")
async def seed_data(force: bool = False):
    # Check if data already exists
    existing_count = await db.employees.count_documents({})
    
    if existing_count > 0 and not force:
        return {"message": "Data already exists. Use force=true to reset all data."}
    
    if force:
        # Clear existing data only if forced
        await db.companies.delete_many({})
        await db.employees.delete_many({})
        await db.roles.delete_many({})
        await db.shift_types.delete_many({})
        await db.attendance.delete_many({})
        await db.leave_records.delete_many({})
        await db.shift_calendar.delete_many({})
        await db.tasks.delete_many({})
        await db.yemek_ucreti.delete_many({})
        await db.avans.delete_many({})
        await db.stok_birim.delete_many({})
        await db.stok_kategori.delete_many({})
        await db.stok_urun.delete_many({})
        await db.stok_sayim.delete_many({})
    
    # Seed companies
    companies = [
        {"id": 1, "name": "Demo Şirket", "domain": "demo.com", "created_at": datetime.now(timezone.utc).isoformat()}
    ]
    await db.companies.insert_many(companies)
    
    # Seed roles
    roles = [
        {
            "id": "admin",
            "name": "Yönetici",
            "permissions": {
                "view_dashboard": True,
                "view_tasks": True,
                "assign_tasks": True,
                "rate_tasks": True,
                "manage_shifts": True,
                "manage_leave": True,
                "view_salary": True,
                "manage_roles": True,
                "manage_shifts_types": True,
                "edit_employees": True,
                "can_view_stock": True,
                "can_add_stock_unit": True,
                "can_delete_stock_unit": True,
                "can_add_stock_product": True,
                "can_edit_stock_product": True,
                "can_delete_stock_product": True,
                "can_perform_stock_count": True,
                "can_manage_categories": True
            }
        },
        {
            "id": "manager",
            "name": "Müdür",
            "permissions": {
                "view_dashboard": True,
                "view_tasks": True,
                "assign_tasks": True,
                "rate_tasks": True,
                "manage_shifts": True,
                "manage_leave": True,
                "view_salary": True,
                "manage_roles": False,
                "manage_shifts_types": True,
                "edit_employees": True,
                "can_view_stock": True,
                "can_add_stock_unit": True,
                "can_delete_stock_unit": False,
                "can_add_stock_product": True,
                "can_edit_stock_product": True,
                "can_delete_stock_product": False,
                "can_perform_stock_count": True,
                "can_manage_categories": True
            }
        },
        {
            "id": "employee",
            "name": "Çalışan",
            "permissions": {
                "view_dashboard": True,
                "view_tasks": True,
                "assign_tasks": False,
                "rate_tasks": False,
                "manage_shifts": False,
                "manage_leave": False,
                "view_salary": True,
                "manage_roles": False,
                "manage_shifts_types": False,
                "edit_employees": False,
                "can_view_stock": True,
                "can_add_stock_unit": False,
                "can_delete_stock_unit": False,
                "can_add_stock_product": False,
                "can_edit_stock_product": False,
                "can_delete_stock_product": False,
                "can_perform_stock_count": False,
                "can_manage_categories": False
            }
        }
    ]
    await db.roles.insert_many(roles)
    
    # Seed shift types
    shift_types = [
        {"id": "sabah", "name": "Sabah", "start": "08:00", "end": "16:00", "color": "#3B82F6"},
        {"id": "ogle_sonra", "name": "Öğleden Sonra", "start": "16:00", "end": "00:00", "color": "#EF4444"},
        {"id": "gece", "name": "Gece", "start": "00:00", "end": "08:00", "color": "#6B7280"},
        {"id": "tam_gun", "name": "Tam Gün", "start": "09:00", "end": "18:00", "color": "#10B981"}
    ]
    await db.shift_types.insert_many(shift_types)
    
    # Seed employees with hashed passwords
    employees = [
        {
            "id": 1,
            "company_id": 1,
            "ad": "Admin",
            "soyad": "User",
            "pozisyon": "Sistem Yöneticisi",
            "maas_tabani": 50000,
            "rol": "admin",
            "email": "admin@example.com",
            "employee_id": "1000",
            "password": bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        },
        {
            "id": 2,
            "company_id": 1,
            "ad": "Mehmet",
            "soyad": "Yılmaz",
            "pozisyon": "Yazılım Geliştirici",
            "maas_tabani": 35000,
            "rol": "employee",
            "email": "mehmet@demo.com",
            "employee_id": "1001",
            "password": bcrypt.hashpw("mehmet123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        },
        {
            "id": 3,
            "company_id": 1,
            "ad": "Zeynep",
            "soyad": "Demir",
            "pozisyon": "Proje Yöneticisi",
            "maas_tabani": 45000,
            "rol": "manager",
            "email": "zeynep@demo.com",
            "employee_id": "1002",
            "password": bcrypt.hashpw("zeynep123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        },
        {
            "id": 4,
            "company_id": 1,
            "ad": "Ayşe",
            "soyad": "Kaya",
            "pozisyon": "İK Uzmanı",
            "maas_tabani": 30000,
            "rol": "employee",
            "email": "ayse@demo.com",
            "employee_id": "1003",
            "password": bcrypt.hashpw("ayse123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        },
        {
            "id": 5,
            "company_id": 1,
            "ad": "Ali",
            "soyad": "Öztürk",
            "pozisyon": "Satış Müdürü",
            "maas_tabani": 42000,
            "rol": "manager",
            "email": "ali@demo.com",
            "employee_id": "1004",
            "password": bcrypt.hashpw("ali123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        },
        {
            "id": 6,
            "company_id": 1,
            "ad": "Arda",
            "soyad": "Yıldız",
            "pozisyon": "Müdür Yardımcısı",
            "maas_tabani": 38000,
            "rol": "manager",
            "email": "arda@demo.com",
            "employee_id": "2001",
            "password": bcrypt.hashpw("arda2024".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        }
    ]
    await db.employees.insert_many(employees)
    
    # Seed attendance records (including for Arda)
    today = datetime.now(timezone.utc).date().isoformat()
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).date().isoformat()
    two_days_ago = (datetime.now(timezone.utc) - timedelta(days=2)).date().isoformat()
    three_days_ago = (datetime.now(timezone.utc) - timedelta(days=3)).date().isoformat()
    four_days_ago = (datetime.now(timezone.utc) - timedelta(days=4)).date().isoformat()
    
    attendance = [
        {
            "id": 1,
            "company_id": 1,
            "employee_id": "1001",
            "ad": "Mehmet",
            "soyad": "Yılmaz",
            "tarih": today,
            "giris_saati": datetime.now(timezone.utc).replace(hour=9, minute=0).isoformat(),
            "cikis_saati": None,
            "calisilan_saat": 0,
            "status": "giris"
        },
        {
            "id": 2,
            "company_id": 1,
            "employee_id": "2001",
            "ad": "Arda",
            "soyad": "Yıldız",
            "tarih": today,
            "giris_saati": datetime.now(timezone.utc).replace(hour=8, minute=30).isoformat(),
            "cikis_saati": None,
            "calisilan_saat": 0,
            "status": "giris"
        },
        {
            "id": 3,
            "company_id": 1,
            "employee_id": "2001",
            "ad": "Arda",
            "soyad": "Yıldız",
            "tarih": yesterday,
            "giris_saati": (datetime.now(timezone.utc) - timedelta(days=1)).replace(hour=9, minute=0).isoformat(),
            "cikis_saati": (datetime.now(timezone.utc) - timedelta(days=1)).replace(hour=18, minute=30).isoformat(),
            "calisilan_saat": 9.5,
            "status": "cikis"
        },
        {
            "id": 4,
            "company_id": 1,
            "employee_id": "2001",
            "ad": "Arda",
            "soyad": "Yıldız",
            "tarih": two_days_ago,
            "giris_saati": (datetime.now(timezone.utc) - timedelta(days=2)).replace(hour=9, minute=0).isoformat(),
            "cikis_saati": (datetime.now(timezone.utc) - timedelta(days=2)).replace(hour=18, minute=0).isoformat(),
            "calisilan_saat": 9.0,
            "status": "cikis"
        },
        {
            "id": 5,
            "company_id": 1,
            "employee_id": "2001",
            "ad": "Arda",
            "soyad": "Yıldız",
            "tarih": three_days_ago,
            "giris_saati": (datetime.now(timezone.utc) - timedelta(days=3)).replace(hour=9, minute=0).isoformat(),
            "cikis_saati": (datetime.now(timezone.utc) - timedelta(days=3)).replace(hour=19, minute=0).isoformat(),
            "calisilan_saat": 10.0,
            "status": "cikis"
        },
        {
            "id": 6,
            "company_id": 1,
            "employee_id": "2001",
            "ad": "Arda",
            "soyad": "Yıldız",
            "tarih": four_days_ago,
            "giris_saati": (datetime.now(timezone.utc) - timedelta(days=4)).replace(hour=9, minute=0).isoformat(),
            "cikis_saati": (datetime.now(timezone.utc) - timedelta(days=4)).replace(hour=18, minute=15).isoformat(),
            "calisilan_saat": 9.25,
            "status": "cikis"
        }
    ]
    await db.attendance.insert_many(attendance)
    
    # Seed yemek ücretleri
    yemek_ucretleri = [
        {"id": 1, "company_id": 1, "employee_id": 1, "gunluk_ucret": 150},
        {"id": 2, "company_id": 1, "employee_id": 2, "gunluk_ucret": 150},
        {"id": 3, "company_id": 1, "employee_id": 3, "gunluk_ucret": 180},
        {"id": 4, "company_id": 1, "employee_id": 4, "gunluk_ucret": 150},
        {"id": 5, "company_id": 1, "employee_id": 5, "gunluk_ucret": 180},
        {"id": 6, "company_id": 1, "employee_id": 6, "gunluk_ucret": 202}
    ]
    await db.yemek_ucreti.insert_many(yemek_ucretleri)
    
    # Seed avans kayıtları (Arda için örnek)
    current_month = datetime.now(timezone.utc).date().isoformat()[:7]
    avans_list = [
        {"id": 1, "company_id": 1, "employee_id": 6, "miktar": 5000, "tarih": f"{current_month}-05", "aciklama": "Kira ödemesi için avans", "olusturan_id": 1},
        {"id": 2, "company_id": 1, "employee_id": 6, "miktar": 2000, "tarih": f"{current_month}-15", "aciklama": "Acil ihtiyaç", "olusturan_id": 1},
        {"id": 3, "company_id": 1, "employee_id": 2, "miktar": 1500, "tarih": f"{current_month}-10", "aciklama": "Avans talebi", "olusturan_id": 1}
    ]
    await db.avans.insert_many(avans_list)
    
    # Seed some leave records
    leave_records = [
        {"id": 1, "company_id": 1, "employee_id": 2, "tarih": "2025-02-15", "leave_type": "izin", "notlar": "Kişisel işler"},
        {"id": 2, "company_id": 1, "employee_id": 4, "tarih": "2025-02-20", "leave_type": "hastalik", "notlar": "Grip"},
        {"id": 3, "company_id": 1, "employee_id": 4, "tarih": "2025-10-25", "leave_type": "izin", "notlar": "Haftalık izin"}
    ]
    await db.leave_records.insert_many(leave_records)
    
    # Seed shift calendar (Ayşe Kaya için örnek - 20-24 Ekim öğleden sonra, 26 Ekim sabah)
    shift_calendar_records = [
        {"id": 1, "company_id": 1, "employee_id": 1, "tarih": "2025-10-20", "shift_type": "sabah"},
        {"id": 2, "company_id": 1, "employee_id": 4, "tarih": "2025-10-20", "shift_type": "ogle_sonra"},
        {"id": 3, "company_id": 1, "employee_id": 3, "tarih": "2025-10-20", "shift_type": "ogle_sonra"},
        {"id": 4, "company_id": 1, "employee_id": 4, "tarih": "2025-10-21", "shift_type": "ogle_sonra"},
        {"id": 5, "company_id": 1, "employee_id": 2, "tarih": "2025-10-21", "shift_type": "ogle_sonra"},
        {"id": 6, "company_id": 1, "employee_id": 4, "tarih": "2025-10-22", "shift_type": "ogle_sonra"},
        {"id": 7, "company_id": 1, "employee_id": 5, "tarih": "2025-10-22", "shift_type": "ogle_sonra"},
        {"id": 8, "company_id": 1, "employee_id": 4, "tarih": "2025-10-23", "shift_type": "ogle_sonra"},
        {"id": 9, "company_id": 1, "employee_id": 1, "tarih": "2025-10-23", "shift_type": "ogle_sonra"},
        {"id": 10, "company_id": 1, "employee_id": 4, "tarih": "2025-10-24", "shift_type": "ogle_sonra"},
        {"id": 11, "company_id": 1, "employee_id": 6, "tarih": "2025-10-24", "shift_type": "ogle_sonra"},
        {"id": 12, "company_id": 1, "employee_id": 4, "tarih": "2025-10-26", "shift_type": "sabah"},
        {"id": 13, "company_id": 1, "employee_id": 2, "tarih": "2025-10-26", "shift_type": "sabah"},
        {"id": 14, "company_id": 1, "employee_id": 3, "tarih": "2025-10-26", "shift_type": "sabah"}
    ]
    await db.shift_calendar.insert_many(shift_calendar_records)
    
    # Seed some tasks
    tasks = [
        {
            "id": 1,
            "company_id": 1,
            "baslik": "Website Tasarımı",
            "aciklama": "Yeni kurumsal website tasarımı yapılacak",
            "atanan_personel_ids": [2],
            "olusturan_id": 1,
            "durum": "devam_ediyor",
            "puan": None,
            "olusturma_tarihi": (datetime.now(timezone.utc) - timedelta(days=2)).isoformat(),
            "tamamlanma_tarihi": None,
            "tekrarlayan": False,
            "tekrar_periyot": None
        },
        {
            "id": 2,
            "company_id": 1,
            "baslik": "Yeni Menü Hazırlığı",
            "aciklama": "Bahar menüsü için yeni yemekler geliştirilecek",
            "atanan_personel_ids": [3],
            "olusturan_id": 1,
            "durum": "beklemede",
            "puan": None,
            "olusturma_tarihi": datetime.now(timezone.utc).isoformat(),
            "tamamlanma_tarihi": None,
            "tekrarlayan": False,
            "tekrar_periyot": None
        },
        {
            "id": 3,
            "company_id": 1,
            "baslik": "Aylık Stok Sayımı",
            "aciklama": "Depo ve mutfak stok sayımı yapılacak, eksik malzemeler listelenecek",
            "atanan_personel_ids": [3, 6],
            "olusturan_id": 1,
            "durum": "beklemede",
            "puan": None,
            "olusturma_tarihi": datetime.now(timezone.utc).isoformat(),
            "tamamlanma_tarihi": None,
            "tekrarlayan": True,
            "tekrar_periyot": "özel",
            "tekrar_sayi": 1,
            "tekrar_birim": "ay"
        }
    ]
    await db.tasks.insert_many(tasks)
    
    # Seed stok kategorileri
    stok_kategorileri = [
        {"id": 1, "company_id": 1, "ad": "Malzeme", "renk": "#9333EA"},  # purple
        {"id": 2, "company_id": 1, "ad": "İçecek", "renk": "#3B82F6"},   # blue
        {"id": 3, "company_id": 1, "ad": "Temizlik", "renk": "#10B981"}, # green
        {"id": 4, "company_id": 1, "ad": "Diğer", "renk": "#6B7280"}     # gray
    ]
    await db.stok_kategori.insert_many(stok_kategorileri)
    
    # Seed stok birimleri
    stok_birimleri = [
        {"id": 1, "company_id": 1, "ad": "Kilogram", "kisaltma": "kg"},
        {"id": 2, "company_id": 1, "ad": "Gram", "kisaltma": "gr"},
        {"id": 3, "company_id": 1, "ad": "Litre", "kisaltma": "lt"},
        {"id": 4, "company_id": 1, "ad": "Adet", "kisaltma": "adet"},
        {"id": 5, "company_id": 1, "ad": "Paket", "kisaltma": "paket"}
    ]
    await db.stok_birim.insert_many(stok_birimleri)
    
    # Seed stok ürünleri
    stok_urunleri = [
        {"id": 1, "company_id": 1, "ad": "Domates", "birim_id": 1, "kategori_id": 1, "min_stok": 5.0},
        {"id": 2, "company_id": 1, "ad": "Soğan", "birim_id": 1, "kategori_id": 1, "min_stok": 3.0},
        {"id": 3, "company_id": 1, "ad": "Coca Cola", "birim_id": 3, "kategori_id": 2, "min_stok": 10.0},
        {"id": 4, "company_id": 1, "ad": "Ekmek", "birim_id": 4, "kategori_id": 1, "min_stok": 20.0},
        {"id": 5, "company_id": 1, "ad": "Süt", "birim_id": 3, "kategori_id": 1, "min_stok": 5.0},
        {"id": 6, "company_id": 1, "ad": "Yumurta", "birim_id": 4, "kategori_id": 1, "min_stok": 30.0},
        {"id": 7, "company_id": 1, "ad": "Çay", "birim_id": 5, "kategori_id": 2, "min_stok": 2.0}
    ]
    await db.stok_urun.insert_many(stok_urunleri)
    
    # Seed stok sayımları
    today_str = datetime.now(timezone.utc).date().isoformat()
    stok_sayimlari = [
        {"id": 1, "company_id": 1, "urun_id": 1, "miktar": 8.5, "tarih": today_str, "sayim_yapan_id": 3, "notlar": "İlk sayım"},
        {"id": 2, "company_id": 1, "urun_id": 2, "miktar": 4.2, "tarih": today_str, "sayim_yapan_id": 3, "notlar": "İlk sayım"},
        {"id": 3, "company_id": 1, "urun_id": 3, "miktar": 15.0, "tarih": today_str, "sayim_yapan_id": 6, "notlar": "İlk sayım"},
        {"id": 4, "company_id": 1, "urun_id": 4, "miktar": 25.0, "tarih": today_str, "sayim_yapan_id": 6, "notlar": "İlk sayım"},
        {"id": 5, "company_id": 1, "urun_id": 5, "miktar": 6.5, "tarih": today_str, "sayim_yapan_id": 3, "notlar": "İlk sayım"},
        {"id": 6, "company_id": 1, "urun_id": 6, "miktar": 45.0, "tarih": today_str, "sayim_yapan_id": 3, "notlar": "İlk sayım"},
        {"id": 7, "company_id": 1, "urun_id": 7, "miktar": 1.0, "tarih": today_str, "sayim_yapan_id": 6, "notlar": "Kritik seviye - yenilenmeli"}
    ]
    await db.stok_sayim.insert_many(stok_sayimlari)
    
    return {"message": "Demo veriler başarıyla yüklendi"}

# Include router AFTER middleware
# NOTE: router inclusion moved to bottom so all @api_router routes defined below are registered
# app.include_router(api_router)


# ==================== SALARY / AVANS / YEMEK ENDPOINTS ====================


@api_router.get("/avans", response_model=List[Avans])
async def get_avans(company_id: int = 1):
    avans = await db.avans.find({"company_id": company_id}).to_list(None)
    return avans


@api_router.post("/avans", response_model=Avans)
async def create_avans(avans_data: AvansCreate, olusturan_id: int = 1):
    next_id = await get_next_id("avans")
    new_avans = {
        "id": next_id,
        "company_id": avans_data.company_id or 1,
        "employee_id": avans_data.employee_id,
        "miktar": avans_data.miktar,
        "tarih": avans_data.tarih if hasattr(avans_data, 'tarih') else datetime.now(timezone.utc).date().isoformat(),
        "aciklama": avans_data.aciklama if hasattr(avans_data, 'aciklama') else "",
        "olusturan_id": olusturan_id
    }
    await db.avans.insert_one(new_avans)
    return new_avans


@api_router.delete("/avans/{avans_id}")
async def delete_avans(avans_id: int):
    result = await db.avans.delete_one({"id": avans_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Avans kaydı bulunamadı")
    return {"message": "Avans silindi"}


@api_router.get("/yemek-ucreti")
async def get_yemek_ucretleri(company_id: int = 1):
    y = await db.yemek_ucreti.find({"company_id": company_id}).to_list(None)
    return y


@api_router.post("/yemek-ucreti")
async def set_yemek_ucreti(employee_id: int, gunluk_ucret: float, company_id: int = 1):
    # Upsert daily meal allowance for an employee (employee_id here is numeric employee.id)
    existing = await db.yemek_ucreti.find_one({"company_id": company_id, "employee_id": int(employee_id)})
    if existing:
        await db.yemek_ucreti.update_one({"id": existing["id"]}, {"$set": {"gunluk_ucret": float(gunluk_ucret)}})
        updated = await db.yemek_ucreti.find_one({"id": existing["id"]})
        return updated

    next_id = await get_next_id("yemek_ucreti")
    new_y = {"id": next_id, "company_id": company_id, "employee_id": int(employee_id), "gunluk_ucret": float(gunluk_ucret)}
    await db.yemek_ucreti.insert_one(new_y)
    return new_y


@api_router.get("/salary-all/{month}")
async def salary_all(month: str):
    """Return aggregated salary records for given month (format: YYYY-MM)."""
    # Validate month format loosely
    if not month or len(month) < 7:
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")

    # 1. Fetch all data in parallel
    employees_cursor = db.employees.find({})
    attendance_cursor = db.attendance.find({"tarih": {"$regex": f"^{month}"}})
    avans_cursor = db.avans.find({"tarih": {"$regex": f"^{month}"}})
    yemek_ucreti_cursor = db.yemek_ucreti.find({})

    employees, attendance_records_all, avans_records_all, yemek_ucretleri_all = await asyncio.gather(
        employees_cursor.to_list(None),
        attendance_cursor.to_list(None),
        avans_cursor.to_list(None),
        yemek_ucreti_cursor.to_list(None),
    )

    # 2. Process data in memory
    attendance_by_employee = {}
    for record in attendance_records_all:
        emp_id = record.get("employee_id")
        if emp_id not in attendance_by_employee:
            attendance_by_employee[emp_id] = []
        attendance_by_employee[emp_id].append(record)

    avans_by_employee = {}
    for record in avans_records_all:
        emp_id = record.get("employee_id")
        if emp_id not in avans_by_employee:
            avans_by_employee[emp_id] = []
        avans_by_employee[emp_id].append(record)

    yemek_ucreti_by_employee = {item['employee_id']: item for item in yemek_ucretleri_all}


    results = []
    for emp in employees:
        # basic fields
        temel = float(emp.get("maas_tabani", 0) or 0)
        gunluk = round(temel / 30.0, 2)
        saatlik = round(gunluk / 9.0, 2)  # assume 9h workday

        emp_id_str = emp.get("employee_id")
        emp_id_int = emp.get("id")

        # Get attendance for this employee from the pre-fetched data
        attendance_records = attendance_by_employee.get(emp_id_str, [])
        if not attendance_records and emp_id_int is not None:
             attendance_records = attendance_by_employee.get(str(emp_id_int), [])


        # Count worked days and sum hours
        calisilan_gun = 0
        calisilan_saat = 0.0
        for a in attendance_records:
            try:
                cs = float(a.get("calisilan_saat", 0) or 0)
            except Exception:
                cs = 0.0
            if cs > 0 or a.get("status") == "cikis":
                calisilan_gun += 1
                calisilan_saat += cs

        hakedilen = round(saatlik * calisilan_saat, 2)

        # Get meal allowance from pre-fetched data
        yemek_doc = yemek_ucreti_by_employee.get(emp_id_int)
        gunluk_yemek = float(yemek_doc.get("gunluk_ucret", 0)) if yemek_doc else 0.0
        toplam_yemek = round(gunluk_yemek * calisilan_gun, 2)

        # Get advances from pre-fetched data
        avans_records = avans_by_employee.get(emp_id_int, [])
        toplam_avans = round(sum([float(a.get("miktar", 0) or 0) for a in avans_records]), 2)

        toplam = round(hakedilen + toplam_yemek - toplam_avans, 2)

        record = {
            "employee_id": emp_id_int,
            "employee_unique_id": emp_id_str,
            "ad": emp.get("ad"),
            "soyad": emp.get("soyad"),
            "pozisyon": emp.get("pozisyon", ""),
            "temel_maas": temel,
            "gunluk_maas": gunluk,
            "saatlik_maas": saatlik,
            "calisilan_gun": calisilan_gun,
            "calisilan_saat": round(calisilan_saat, 2),
            "hakedilen_maas": hakedilen,
            "gunluk_yemek_ucreti": gunluk_yemek,
            "toplam_yemek": toplam_yemek,
            "toplam_avans": toplam_avans,
            "toplam_maas": toplam,
            "ay": month
        }

        results.append(record)

    return results


def _workbook_from_dicts(rows, headers=None, sheet_name="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if not rows:
        if headers:
            ws.append(headers)
        return wb
    if not headers:
        # derive headers from first row
        headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    return wb


@api_router.get("/salary-all/{month}/xlsx")
async def salary_all_xlsx(month: str):
    # reuse salary_all logic (call the function body or duplicate minimal logic)
    data = await salary_all(month)
    # create workbook
    headers = ["employee_id", "employee_unique_id", "ad", "soyad", "pozisyon", "temel_maas", "gunluk_maas", "saatlik_maas", "calisilan_gun", "calisilan_saat", "hakedilen_maas", "gunluk_yemek_ucreti", "toplam_yemek", "toplam_avans", "toplam_maas", "ay"]
    wb = _workbook_from_dicts(data, headers=headers, sheet_name=f"Salary_{month}")
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    fname = f"salary_{month}.xlsx"
    return StreamingResponse(stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename=\"{fname}\""})


@api_router.get("/stok-export")
async def stok_export(company_id: int = 1):
    # export stock products and last known stock (if any)
    urunler = await db.stok_urun.find({"company_id": company_id}).to_list(None)
    # try to get last sayim quantities
    results = []
    for u in urunler:
        # lookup latest sayim for this product
        sayim = await db.stok_sayim.find({"urun_id": u.get("id")}).sort([("tarih", -1)]).to_list(1)
        miktar = sayim[0].get("miktar") if sayim else ""
        results.append({
            "id": u.get("id"),
            "ad": u.get("ad"),
            "birim_id": u.get("birim_id"),
            "kategori_id": u.get("kategori_id"),
            "min_stok": u.get("min_stok", 0),
            "mevcut_miktar": miktar,
        })
    wb = _workbook_from_dicts(results, headers=["id", "ad", "birim_id", "kategori_id", "min_stok", "mevcut_miktar"], sheet_name="Stok")
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": "attachment; filename=\"stok_export.xlsx\""})


@api_router.get("/stok-template")
async def stok_template():
    # return a sample template XLSX for bulk stock upload
    sample = [
        {"ad": "Domates", "birim_id": 1, "kategori_id": 1, "min_stok": 10, "mevcut_miktar": 50},
        {"ad": "Pirinç", "birim_id": 2, "kategori_id": 2, "min_stok": 5, "mevcut_miktar": 200},
    ]
    wb = _workbook_from_dicts(sample, headers=["ad", "birim_id", "kategori_id", "min_stok", "mevcut_miktar"], sheet_name="Template")
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": "attachment; filename=\"stok_template.xlsx\""})


@api_router.post("/stok-import")
async def stok_import(file: UploadFile = File(...), company_id: int = 1):
    # parse uploaded XLSX and insert/update stok_urun
    try:
        content = await file.read()
        stream = io.BytesIO(content)
        wb = openpyxl.load_workbook(stream)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise HTTPException(status_code=400, detail="Dosya boş veya başlık yok")
        headers = [str(h).strip() for h in rows[0]]
        created = 0
        updated = 0
        for r in rows[1:]:
            data = {headers[i]: r[i] for i in range(len(headers))}
            # expect at least 'ad' field
            if not data.get('ad'):
                continue
            # try to match by name
            existing = await db.stok_urun.find_one({"company_id": company_id, "ad": data.get('ad')})
            doc = {
                "company_id": company_id,
                "ad": data.get('ad'),
                "birim_id": int(data.get('birim_id')) if data.get('birim_id') else None,
                "kategori_id": int(data.get('kategori_id')) if data.get('kategori_id') else None,
                "min_stok": float(data.get('min_stok') or 0)
            }
            if existing:
                await db.stok_urun.update_one({"id": existing['id']}, {"$set": doc})
                updated += 1
            else:
                next_id = await get_next_id('stok_urun')
                doc['id'] = next_id
                await db.stok_urun.insert_one(doc)
                created += 1
        return {"created": created, "updated": updated}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Debug endpoint to validate Sentry and logging in production.
# Usage:
#  - POST /api/debug/sentry-test            -> sends a Sentry message (if configured) and returns 200
#  - POST /api/debug/sentry-test?raise_exc=true -> triggers an unhandled exception (500) to force Sentry to capture an exception
@api_router.post("/debug/sentry-test")
async def sentry_test(raise_exc: bool = False):
    """Send a test message to Sentry (if available). If raise_exc=true, raise an exception to produce a 500 and ensure Sentry captures it.

    This endpoint is intended for manual testing after you set SENTRY_DSN in your Render environment.
    """
    try:
        import sentry_sdk
        # Capture an informational message so Sentry shows an event even if we don't raise
        sentry_sdk.capture_message("Mevcut App - Sentry test message from /api/debug/sentry-test")
        logger.info("Sentry test message captured (if Sentry configured)")
    except Exception:
        logger.exception("Sentry SDK not configured or capture failed")

    if raise_exc:
        # This will produce a 500 response and, if Sentry is configured, an exception event
        raise Exception("Sentry test exception triggered by /api/debug/sentry-test")

    return {"success": True, "sentry_message_sent": True, "raise_exception": raise_exc}


# Include API router after all routes have been declared so every @api_router
# route is registered. Previously include_router was called earlier which
# caused routes defined after that call to return 404 (e.g. salary endpoints).
# NOTE: router will be included at the end of this file after subscription/webhook handlers

# Import optional modules that register additional routes (pos, etc.).
try:
    # import side-effect modules that call api_router decorators
    from . import pos  # noqa: F401
except Exception:
    try:
        import pos  # fallback if package context differs
    except Exception:
        logger.info("Optional module 'pos' not available or failed to import at startup")


# -------------------- Subscription / Payment endpoints --------------------


@api_router.get("/subscription/company/{company_id}")
async def get_subscription(company_id: int):
    """Return subscription status for a given company.

    Subscriptions are maintained per company (tenant). Employees belong to a company
    and access company-level data. This endpoint returns the subscription record for
    the provided company id.
    """
    sub = await db.subscriptions.find_one({"company_id": int(company_id)})
    if not sub:
        return {"company_id": int(company_id), "status": "none"}

    # sanitize and expose useful billing fields for frontend
    plan = sub.get("plan") or {}
    # plan may be stored as string (price id) or dict with details
    plan_obj = plan if isinstance(plan, dict) else {"id": plan}

    return {
        "company_id": int(company_id),
        "status": sub.get("status", "unknown"),
        "plan": plan_obj,
        "stripe_subscription_id": sub.get("stripe_subscription_id"),
        "billing_email": sub.get("billing_email"),
        "current_period_end": sub.get("current_period_end"),
        "created_at": sub.get("created_at"),
        "updated_at": sub.get("updated_at")
    }


@api_router.post("/create-checkout-session")
async def create_checkout_session(payload: dict):
    """Create a Stripe Checkout Session for a subscription purchase.

    Expected JSON payload:
      { "employee_id": 6, "price_id": "price_...", "success_url": "https://.../success", "cancel_url": "https://.../cancel" }

    If STRIPE_SECRET is not configured we'll fallback to a mock path that
    persist a subscription record with status 'active' for testing.
    """
    company_id = int(payload.get("company_id") or 0)
    price_id = payload.get("price_id")
    success_url = payload.get("success_url") or "https://example.com/success"
    cancel_url = payload.get("cancel_url") or "https://example.com/cancel"

    stripe_key = os.environ.get("STRIPE_SECRET")
    if not stripe_key:
        # Mock flow: create a subscription record locally and return a mock url
        # Support optional billing_email and plan_name in payload for better demo UX
        billing_email = payload.get("billing_email") or payload.get("customer_email")
        plan_name = payload.get("plan_name") or ("Aylık Plan" if (price_id or '').lower().find('month') >= 0 else "Yıllık Plan")
        # Compute a mock period end date (30 days for monthly, 365 for yearly)
        period_days = 30 if 'month' in (price_id or '').lower() else 365
        current_period_end = (datetime.now(timezone.utc) + timedelta(days=period_days)).date().isoformat()

        sub_doc = {
            "company_id": company_id,
            "status": "active",
            "plan": {
                "id": price_id or "mock_monthly",
                "name": plan_name,
                "price_display": ("₺49 / ay" if (price_id or '').lower().find('month') >= 0 else "₺499 / yıl")
            },
            "billing_email": billing_email,
            "current_period_end": current_period_end,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "stripe_subscription_id": None,
        }
        await db.subscriptions.insert_one(sub_doc)
        return {"mock": True, "message": "Subscription created (mock)", "redirect": success_url}

    try:
        stripe.api_key = stripe_key
        # Create a Checkout Session for subscription purchase
        session = stripe.checkout.Session.create(
            mode="subscription",
            line_items=[{"price": price_id, "quantity": 1}],
            success_url=success_url + "?session_id={CHECKOUT_SESSION_ID}",
            cancel_url=cancel_url,
            metadata={"company_id": str(company_id)}
        )
        return {"session_id": session.id, "session_url": session.url}
    except Exception as e:
        logger.exception("Failed to create Stripe checkout session")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/webhook")
async def stripe_webhook(request: Request):
    """Stripe webhook endpoint to receive subscription events.

    Configure STRIPE_WEBHOOK_SECRET in the environment for signature verification.
    """
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature")
    webhook_secret = os.environ.get("STRIPE_WEBHOOK_SECRET")

    if webhook_secret:
        try:
            event = stripe.Webhook.construct_event(payload=payload, sig_header=sig_header, secret=webhook_secret)
        except Exception as e:
            logger.error("Webhook signature verification failed: %s", e)
            raise HTTPException(status_code=400, detail="Invalid webhook signature")
    else:
        # If no webhook secret is configured, attempt to parse the payload (unsafe)
        try:
            event = json.loads(payload.decode('utf-8'))
        except Exception as e:
            logger.error("Failed to parse webhook payload: %s", e)
            raise HTTPException(status_code=400, detail="Invalid payload")

    # Handle relevant event types
    kind = event.get("type") if isinstance(event, dict) else getattr(event, 'type', None)

    try:
        if kind == "checkout.session.completed":
            session = event.get("data", {}).get("object") if isinstance(event, dict) else event.data.object
            company_id = int(session.get("metadata", {}).get("company_id") or 0)
            stripe_sub_id = session.get("subscription")
            # try to collect billing email from session (customer_details in newer API)
            billing_email = None
            if isinstance(session, dict):
                billing_email = session.get("customer_details", {}).get("email") or session.get("customer_email")

            await db.subscriptions.update_one(
                {"company_id": company_id},
                {"$set": {
                    "status": "active",
                    "stripe_subscription_id": stripe_sub_id,
                    "billing_email": billing_email,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }},
                upsert=True
            )
            logger.info("Subscription activated for company %s via checkout.session.completed", company_id)

        elif kind == "invoice.payment_failed":
            invoice = event.get("data", {}).get("object") if isinstance(event, dict) else event.data.object
            sub_id = invoice.get("subscription")
            # mark subscription as past_due or unpaid
            await db.subscriptions.update_one({"stripe_subscription_id": sub_id}, {"$set": {"status": "past_due", "updated_at": datetime.now(timezone.utc).isoformat()}})
            logger.warning("Subscription %s marked past_due due to invoice.payment_failed", sub_id)

        elif kind == "customer.subscription.deleted":
            sub = event.get("data", {}).get("object") if isinstance(event, dict) else event.data.object
            sub_id = sub.get("id")
            await db.subscriptions.update_one({"stripe_subscription_id": sub_id}, {"$set": {"status": "canceled", "updated_at": datetime.now(timezone.utc).isoformat()}})
            logger.info("Subscription %s canceled", sub_id)

    except Exception:
        logger.exception("Error handling webhook event")

    return {"received": True}


# Finally include the API router so all routes are registered
# Import POS routes (module defines endpoints on api_router) if available
try:
    # safe import; module will use the api_router defined above
    from . import pos as _pos_module  # noqa: F401
except Exception:
    try:
        import pos as _pos_module  # noqa: F401
    except Exception:
        logger.info('backend.pos module not imported (may be missing)')

# Staff permissions endpoint for hrAdapter compatibility
@api_router.get("/staff/{staff_id}/permissions")
async def get_staff_permissions(staff_id: int):
    """Return permissions for staff member - currently returns empty dict"""
    # TODO: Implement actual permissions logic when available
    return {}

app.include_router(api_router)

# Debug endpoint to check file structure on Render
@app.get("/debug/files")
async def debug_files():
    try:
        import os
        cwd = os.getcwd()
        result = {
            "cwd": cwd,
            "root_files": os.listdir(".") if os.path.exists(".") else [],
            "frontend_exists": os.path.exists("frontend"),
            "backend_exists": os.path.exists("backend"),
        }
        
        if os.path.exists("frontend"):
            result["frontend_files"] = os.listdir("frontend")
            if os.path.exists("frontend/build"):
                result["frontend_build_files"] = os.listdir("frontend/build")
                if os.path.exists("frontend/build/static/js"):
                    result["js_files"] = os.listdir("frontend/build/static/js")
        
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({"error": str(e)})

# SPA fallback: serve index.html for any non-/api routes so client-side routing works
# This is added after api router registration so API routes keep priority.
try:
    index_file = FRONTEND_BUILD_DIR / "index.html"
    if index_file.exists():
        @app.get("/{full_path:path}")
        async def _spa_fallback(request: Request, full_path: str):
            # let API routes return normally
            if request.url.path.startswith("/api"):
                raise HTTPException(status_code=404)
            return FileResponse(str(index_file))
        logger.info("Registered SPA fallback route serving %s", index_file)
    else:
        logger.info("No frontend index.html found at %s; SPA fallback not registered", FRONTEND_BUILD_DIR)
except Exception:
    logger.exception("Error while registering SPA fallback route")

